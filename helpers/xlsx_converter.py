import openpyxl
import openpyxl.styles
import openpyxl.utils
from bom.structured_bom import TableColors

from openpyxl.worksheet.worksheet import Worksheet

# from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, colors

code_to_color = {
    TableColors.BLANK: openpyxl.styles.colors.Color(rgb='00FFFFFF'),
    TableColors.TITLE: openpyxl.styles.colors.Color(rgb='00DDDDDD'),
    TableColors.SUB_ASSEMBLY: openpyxl.styles.colors.Color(rgb='00D9E1F2'),
    TableColors.LEFT_MATCH: openpyxl.styles.colors.Color(rgb='00E7F8EA'),
    TableColors.LEFT_DIFF: openpyxl.styles.colors.Color(rgb='00FFC7CE'),
    TableColors.LEFT_SHOW: openpyxl.styles.colors.Color(rgb='00EAEAEA'),
    TableColors.RIGHT_MATCH: openpyxl.styles.colors.Color(rgb='00C6EFCE'),
    TableColors.RIGHT_DIFF: openpyxl.styles.colors.Color(rgb='00FFAEB0'),
    TableColors.RIGHT_SHOW: openpyxl.styles.colors.Color(rgb='00DDDDDD'),
}

code_to_fill = {key: openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=color) for key, color in
                code_to_color.items()}

name_to_width = {
    "Item Code": 14,
    "Rev": 4,
    "Description": 40,
    "Qty": 4,
    "#": 6,
}


def report_to_xlsx(report, color_mapping, sheet_name, path=None, workbook=None):
    if workbook is None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
    else:
        worksheet = workbook.create_sheet(sheet_name)

    assert len(color_mapping) == len(report), "%s != %s" % (len(color_mapping), len(report))
    for row_index in range(len(report)):
        row = report[row_index]
        colors = color_mapping[row_index]
        assert len(row) == len(colors), "%s != %s @ %s" % (len(row), len(colors), row_index)

        for col_index in range(len(row)):
            cell_value = row[col_index]
            color_code = colors[col_index]

            cell = worksheet.cell(row=row_index + 1, column=col_index + 1)
            cell.value = cell_value
            cell.fill = code_to_fill[color_code]


    widths = []
    label_row = report[1]
    for col_index in range(len(label_row)):
        label_key = label_row[col_index]
        width = name_to_width.get(label_key, 10)
        widths.append(width)

    for index, width in enumerate(widths):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(index + 1)].width = width

    if path is not None:
        workbook.save(path)

    return workbook
